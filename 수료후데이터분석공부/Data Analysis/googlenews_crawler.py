# 'Google' 접속

from selenium import webdriver
import time
from tqdm import tqdm_notebook
import datetime
import pandas as pd

from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import smtplib
import os

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.styles import Border, Color, PatternFill
from openpyxl import load_workbook


print('Google 기사 데이터 수집 시작')

if __name__ == '__main__' :
    path='./chromedriver.exe'
    driver = webdriver.Chrome(path)
    url = 'https://www.google.com/'
    driver.get(url)


    now = datetime.datetime.now()
    nowDate = now.strftime('%Y-%m-%d')

    print('오늘 날짜 : ', nowDate)
    print('외국기사_구글뉴스_크롤링 시작합니다.')


    time.sleep(2)


    # 뉴스 검색어 입력

    # 검색어를 입력하세요.
    # keyword = input('검색어를 입력하세요: ')

    keywords = ['Fiber Optics', '5G', 'LTE', 'fixed broadband', 'MVNO', 'mobile communication', 'data center', 'OTT', '5G coverage map']



    keyword_input = driver.find_element_by_tag_name('input.gLFyf')
    keyword_input.send_keys(keywords[0])

    driver.find_element_by_tag_name('input.gNO89b').submit()

    time.sleep(2)

    # '뉴스' 컬럼 접속
    # '더보기' 열기
    driver.find_element_by_css_selector('div.hdtb-dd-b').click()
    time.sleep(2)

    view_more = driver.find_elements_by_css_selector('div.gTMtLb > div.EwsJzb > g-menu.cF4V5c > g-menu-item')
    columns = driver.find_elements_by_tag_name('div.hdtb-mitem > a.hide-focus-ring')

    for element in view_more:
        columns.append(element)

    for column in columns:
        if column.text == '뉴스':
            column.send_keys('\n')
            break
        

    time.sleep(3)




    # '영어'로 언어 변경

    setting = driver.find_element_by_tag_name('a.hdtb-dd-b')
    setting.click()

    setting_list = driver.find_elements_by_tag_name('div.gTMtLb > div.sAKBe >g-menu > g-menu-item')
    setting_list[1].click()

    time.sleep(2)

    English = driver.find_element_by_tag_name('div#langtop > div > div#langten')
    English.click()

    lang_save = driver.find_element_by_tag_name('div#form-buttons > div')
    lang_save.click()

    time.sleep(3)


    # alert창 처리
    result = driver.switch_to_alert()
    result.accept()

    time.sleep(3)



    # 뉴스 작성 기간 '최근 1일' 지정

    tools = driver.find_element_by_tag_name('div#hdtb-tls')
    tools.click()

    time.sleep(3)

    period = driver.find_elements_by_tag_name('div.mn-hd-txt')
    period[0].click()

    time.sleep(3)

    driver.find_elements_by_tag_name('div#lb > div > g-menu > g-menu-item')[2].click()

    time.sleep(3)

    period = driver.find_elements_by_tag_name('div.mn-hd-txt')
    period[1].click()

    time.sleep(3)
    driver.find_elements_by_tag_name('div#lb > div > g-menu > g-menu-item')[1].click()





    total_list = []
        
    for keyword in keywords:
        
        url_list = []
        title_list = []
        article_preview = []
        keyword_list = []

        keyword_input = driver.find_element_by_tag_name('input#lst-ib')
        keyword_input.clear()
        keyword_input.send_keys(keyword)
        
        search = driver.find_element_by_tag_name('button#mKlEF')
        search.click()

        time.sleep(2)

        # 뉴스 리스트 크롤링


        print('현재 크롤링 중인 기사 검색어 : ', keyword)
        while 1:
            news_list = driver.find_elements_by_tag_name('div#rso > div')

            for idx in range(len(news_list)):

                news = news_list[idx]

                # 기사 제목
                title = news.find_element_by_tag_name('div.JheGif').text
                title_list.append(title)

                # 기사 본문
                article = news.find_element_by_tag_name('div.yJHHTd > div').text
                article_preview.append(article)


                # 기사 url
                url = news.find_element_by_tag_name('div.dbsr > a').get_attribute('href')
                url_list.append(url)

                keyword_list.append(keyword)

            try :
                driver.find_element_by_tag_name('td.d6cvqb > a#pnnext').click()
                time.sleep(2)

            except:
                break
        print('지금까지 크롤링한 기사 수 : ', len(title_list))

        google_news = pd.DataFrame({'검색어' : keyword_list, '제목' : title_list, '본문': article_preview, 'url': url_list})
        google_news.reset_index(drop=True, inplace=True)
        google_news = google_news.iloc[0:10, :]
        total_list.append(google_news)

    driver.close()

    # xlsx 파일로 저장
    temp_excel = '../temp_excel.xlsx'
    with pd.ExcelWriter(temp_excel) as writer:
        for word, google_news in zip(keywords, total_list):
            google_news.to_excel(writer, sheet_name=word)
    # google_news.to_excel("../Google_뉴스_["+str(nowDate)+"].xlsx",encoding='UTF-8-SIG')

    # 엑셀 서식
    load_excel = load_workbook(filename=temp_excel, data_only=True)    
    for sheet_name in keywords:
        sheet = load_excel[sheet_name]
        # 셀 너비 지정
        sheet.column_dimensions['A'].width = 8
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 55
        sheet.column_dimensions['D'].width = 75
        sheet.column_dimensions['E'].width = 70

        # 셀 가운데 정렬
        for col in [sheet['A'], sheet['B']]:
            for cell in col:
                cell.alignment = Alignment(horizontal = 'center', vertical = 'center')

        # 셀 글꼴 지정(나눔고딕)
        for col in sheet.columns:
            for cell in col:
                cell.font = Font('나눔고딕')
                
        # 첫 행 셀 색상 & Bold 적용
        for col in sheet.columns:
            for idx, cell in enumerate(col):
                if idx == 0:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(fill_type='solid', start_color='dcdcdc', end_color='dcdcdc')
                else:
                    break

    load_excel.save("../Google_뉴스_["+str(nowDate)+"].xlsx")
    os.remove(temp_excel)



    print('')
    print(nowDate)
    print('구글 뉴스 크롤링 완료')



    print('메일 전달 시작')

    # 메일 보내기

    # SMTP 접속을 위한 서버, 계정 설정
    SMTP_SERVER = '페이지 주소 입력'
    SMTP_PORT = 000

    # SMTP로 접속할 서버 정보를 가진 클래스 변수 생성
    smtp = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)

    # 보내는 메일 계정
    SMTP_USER = '메일 아이디 입력'
    SMTP_PASSWORD = '메일 비밀번호 입력'

    # 해당 서버로 로그인
    smtp.login(SMTP_USER, SMTP_PASSWORD)



    msg = MIMEMultipart()
    msg['Subject'] =  '[' + nowDate + ']' + '메일 제목'
    content =  '''메일내용입력''' 

    msg.attach(MIMEText(content,'plain'))



    # 파일 첨부
    attachment = open("../Google_뉴스_["+str(nowDate)+"].xlsx", 'rb')
    part = MIMEBase('application', 'csv')
    part.set_payload((attachment).read())

    # mime은 메세지 말고 다른 이진 파일들을 보내려면 인코딩하여 텍스트로 바꾸어 전달하고 다시 이진파일로 디코딩한다.
    part.add_header('Content-Disposition', 'attachment', filename= 'Google_news_articles_'+str(nowDate)+'.xlsx')
    encoders.encode_base64(part)
    msg.attach(part)

    attachment.close()
    os.remove('../Google_뉴스_["+str(nowDate)+"].xlsx')


    receivers = [ ]
    smtp.sendmail(SMTP_USER, receivers ,msg.as_string())

    print('창을 종료하셔도 좋습니다.')

    

    